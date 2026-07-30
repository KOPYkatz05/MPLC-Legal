"""Transactional, duplicate-safe import of Dynamics missionary rosters."""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from io import BytesIO
import json
import logging
from pathlib import Path
import re
import unicodedata
from uuid import uuid4
import warnings

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

from database.db import SessionLocal
from database.models.missionary import DynamicsRosterImport, Missionary
from database.models.secretary_work import SecretaryTask
from services.onedrive_service import OneDriveService
from services.workflow_service import WorkflowService


CENTRAL_MISSION = "Perú Lima Central Mission (2010429)"
REQUIRED_COLUMNS = {
    "(Do Not Modify) Contact", "(Do Not Modify) Row Checksum",
    "(Do Not Modify) Modified On", "Missionary ID", "Romanized Name",
    "Missionary Status", "Release Date", "Citizenship", "Mission",
    "Home Address", "Mother Name", "Father Name", "In Field Arrival Date",
}
IMPORTABLE_STATUSES = {"In-field", "Delay"}
logger = logging.getLogger(__name__)
UPDATE_FIELDS = (
    "full_name", "nationality", "arrival_date", "release_date",
    "dynamics_contact_id", "dynamics_row_checksum", "dynamics_modified_at",
    "dynamics_status", "home_address", "mother_name", "father_name",
)


class DynamicsRosterError(ValueError):
    pass


def canonical_name(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_name = "".join(char for char in normalized if not unicodedata.combining(char))
    tokens = re.findall(r"[a-z0-9]+", ascii_name.casefold())
    return " ".join(sorted(tokens))


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        for parser in (date.fromisoformat,):
            try:
                return parser(cleaned)
            except ValueError:
                pass
    return None


def _datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return None


def _profile(citizenship):
    value = canonical_name(citizenship)
    return "PERUVIAN_DNI" if value in {"peru", "peruvian"} else "LEGAL"


class DynamicsRosterService:
    def parse(self, content: bytes, filename: str = ""):
        workbook = None
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore", message="Workbook contains no default style.*"
                )
                workbook = load_workbook(
                    BytesIO(content), read_only=True, data_only=True
                )
            if "Missionary - Active" not in workbook.sheetnames:
                raise DynamicsRosterError(
                    "Workbook does not contain the Missionary - Active sheet."
                )
            sheet = workbook["Missionary - Active"]
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, ())]
            missing = sorted(REQUIRED_COLUMNS - set(headers))
            if missing:
                raise DynamicsRosterError("Workbook is missing: " + ", ".join(missing))
            result = []
            for number, raw_values in enumerate(values, start=2):
                source = dict(zip(headers, raw_values))
                code = str(source.get("Missionary ID") or "").strip()
                if not code:
                    continue
                result.append({
                    "row_number": number,
                    "missionary_code": code,
                    "full_name": str(source.get("Romanized Name") or "").strip(),
                    "dynamics_contact_id": str(source.get("(Do Not Modify) Contact") or "").strip(),
                    "dynamics_row_checksum": str(source.get("(Do Not Modify) Row Checksum") or "").strip(),
                    "dynamics_modified_at": _datetime(source.get("(Do Not Modify) Modified On")),
                    "dynamics_status": str(source.get("Missionary Status") or "").strip(),
                    "release_date": _date(source.get("Release Date")),
                    "nationality": str(source.get("Citizenship") or "").strip(),
                    "mission": str(source.get("Mission") or "").strip(),
                    "home_address": str(source.get("Home Address") or "").strip(),
                    "mother_name": str(source.get("Mother Name") or "").strip(),
                    "father_name": str(source.get("Father Name") or "").strip(),
                    "arrival_date": _date(source.get("In Field Arrival Date")),
                })
            return result
        except DynamicsRosterError:
            raise
        except Exception as exc:
            raise DynamicsRosterError(
                "Select a valid Dynamics Missionary Active workbook."
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _safe_row(row):
        return {
            "row": row["row_number"],
            "missionary_code": row["missionary_code"],
            "name": row["full_name"],
            "status": row["dynamics_status"],
            "profile": _profile(row["nationality"]),
        }

    def _classify(self, rows, session):
        missionaries = session.query(Missionary).all()
        by_id = {str(item.missionary_code or "").strip(): item for item in missionaries if item.missionary_code}
        by_name = {}
        for item in missionaries:
            by_name.setdefault(canonical_name(item.full_name), []).append(item)

        creates, changes, unchanged, skipped, invalid, conflicts = [], [], [], [], [], []
        duplicate_codes = {
            code for code, count in Counter(row["missionary_code"] for row in rows).items()
            if count > 1
        }
        accepted = {}
        for row in rows:
            safe = self._safe_row(row)
            code = row["missionary_code"]
            if code in duplicate_codes:
                invalid.append({**safe, "reason": "Duplicate Missionary ID in workbook"})
                continue
            if row["mission"] != CENTRAL_MISSION or row["dynamics_status"] not in IMPORTABLE_STATUSES:
                skipped.append({**safe, "reason": "Outside Central Mission scope"})
                continue
            if not row["full_name"] or not row["arrival_date"] or not row["release_date"]:
                invalid.append({**safe, "reason": "Missing or invalid name, arrival date, or release date"})
                continue

            id_match = by_id.get(code)
            name_matches = by_name.get(canonical_name(row["full_name"]), [])
            if len(name_matches) > 1:
                conflicts.append({
                    **safe, "kind": "MULTIPLE_NAME_MATCHES",
                    "reason": "Multiple existing missionaries have this full name",
                    "existing": [self._existing_summary(item) for item in name_matches],
                    "resolutions": [],
                })
                continue
            name_match = name_matches[0] if name_matches else None
            if id_match and canonical_name(id_match.full_name) != canonical_name(row["full_name"]):
                conflicts.append({
                    **safe, "kind": "ID_CONFLICT",
                    "reason": "Missionary ID matches but the full name differs",
                    "existing": [self._existing_summary(id_match)],
                    "resolutions": ["same", "different"],
                })
                continue
            if name_match and name_match is not id_match and str(name_match.missionary_code or "") != code:
                conflicts.append({
                    **safe, "kind": "POSSIBLE_DUPLICATE",
                    "reason": "Full name matches but the Missionary ID differs",
                    "existing": [self._existing_summary(name_match)],
                    "resolutions": ["same", "different"],
                })
                continue

            match = id_match or name_match
            if match and match.status != "ACTIVE":
                conflicts.append({
                    **safe, "kind": "INACTIVE_MATCH",
                    "reason": f"Matching missionary is {match.status.lower()}",
                    "existing": [self._existing_summary(match)],
                    "resolutions": ["restore", "skip"],
                })
                continue
            if match is None:
                creates.append(safe)
                accepted[row["row_number"]] = ("create", None)
            else:
                fields = self._changed_fields(match, row)
                target = {**safe, "missionary_id": match.id, "fields": fields}
                (changes if fields else unchanged).append(target)
                accepted[row["row_number"]] = ("update", match.id)
        return {
            "creates": creates, "changes": changes, "unchanged": unchanged,
            "skipped": skipped, "invalid": invalid, "conflicts": conflicts,
        }, accepted

    @staticmethod
    def _existing_summary(item):
        return {
            "missionary_id": item.id, "missionary_code": item.missionary_code,
            "name": item.full_name, "status": item.status,
        }

    @staticmethod
    def _changed_fields(record, row):
        changed = [
            field for field in UPDATE_FIELDS
            if getattr(record, field, None) != row[field]
        ]
        profile = _profile(row["nationality"])
        if (record.tracking_profile or "LEGAL") != profile:
            changed.append("tracking_profile")
        target_stage = "DNI" if profile == "PERUVIAN_DNI" else record.current_stage
        if record.current_stage != target_stage:
            changed.append("current_stage")
        return changed

    def preview(self, content: bytes, filename: str = ""):
        digest = sha256(content).hexdigest()
        rows = self.parse(content, filename)
        session = SessionLocal()
        try:
            classified, _accepted = self._classify(rows, session)
            preview_id = str(uuid4())
            summary = {key: len(value) for key, value in classified.items()}
            newest = max(
                (row["dynamics_modified_at"] for row in rows if row["dynamics_modified_at"]),
                default=None,
            )
            audit = DynamicsRosterImport(
                preview_id=preview_id, status="PREVIEW", filename=Path(filename).name,
                filename_timestamp=self._filename_timestamp(filename),
                file_sha256=digest, dynamics_modified_at=newest,
                summary_json=json.dumps(summary),
            )
            session.add(audit)
            session.commit()
            return {
                "preview_id": preview_id, "file_sha256": digest,
                "filename": Path(filename).name,
                "filename_timestamp": audit.filename_timestamp,
                "summary": summary, **classified,
            }
        finally:
            session.close()

    def apply(self, content: bytes, filename: str, preview_id: str, resolutions=None, applying_device=None):
        resolutions = resolutions or {}
        digest = sha256(content).hexdigest()
        rows = self.parse(content, filename)
        session = SessionLocal()
        legal_workflow_ids = []
        try:
            audit = session.query(DynamicsRosterImport).filter_by(preview_id=preview_id).one_or_none()
            if audit is None or audit.status != "PREVIEW":
                raise DynamicsRosterError("This roster preview is missing, stale, or already applied.")
            created_at = audit.created_at
            if created_at is not None:
                now = datetime.now(timezone.utc)
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                if now - created_at > timedelta(hours=24):
                    raise DynamicsRosterError(
                        "This roster preview expired. Preview the workbook again."
                    )
            if audit.file_sha256 != digest:
                raise DynamicsRosterError("The workbook changed after preview. Preview it again.")
            classified, accepted = self._classify(rows, session)
            if classified["invalid"]:
                raise DynamicsRosterError("Correct invalid roster rows before importing.")

            unresolved = []
            for conflict in classified["conflicts"]:
                key = str(conflict["row"])
                resolution = resolutions.get(key)
                if resolution not in conflict["resolutions"]:
                    unresolved.append(conflict)
                    continue
                existing = conflict["existing"][0] if conflict["existing"] else None
                if conflict["kind"] == "ID_CONFLICT":
                    if resolution == "same":
                        accepted[conflict["row"]] = ("update", existing["missionary_id"])
                    # "different" intentionally skips; the conflicting ID is never stolen.
                elif conflict["kind"] == "POSSIBLE_DUPLICATE":
                    if resolution == "same":
                        current = session.get(Missionary, existing["missionary_id"])
                        owner = session.query(Missionary).filter(
                            Missionary.missionary_code == conflict["missionary_code"],
                            Missionary.id != current.id,
                        ).first()
                        if owner:
                            raise DynamicsRosterError(
                                f"Missionary ID {conflict['missionary_code']} is already assigned."
                            )
                        accepted[conflict["row"]] = ("update", current.id)
                    else:
                        accepted[conflict["row"]] = ("create", None)
                elif conflict["kind"] == "INACTIVE_MATCH":
                    if resolution == "restore":
                        accepted[conflict["row"]] = ("restore", existing["missionary_id"])
            if unresolved:
                raise DynamicsRosterError("Resolve every identity conflict before applying.")

            created = updated = unchanged = restored = 0
            affected_ids = []
            restored_folders = []
            row_by_number = {row["row_number"]: row for row in rows}
            for row_number, (action, missionary_id) in accepted.items():
                row = row_by_number[row_number]
                profile = _profile(row["nationality"])
                if action == "create":
                    missionary = Missionary(
                        missionary_code=row["missionary_code"],
                        full_name=row["full_name"],
                        status="ACTIVE", tracking_profile=profile,
                        current_stage="DNI" if profile == "PERUVIAN_DNI" else "INTERPOL",
                        folder_path=str(OneDriveService().create_missionary_folders(row["full_name"])),
                    )
                    session.add(missionary)
                    for field in UPDATE_FIELDS:
                        setattr(missionary, field, row[field])
                    session.flush()
                    affected_ids.append(missionary.id)
                    created += 1
                    if profile == "LEGAL":
                        legal_workflow_ids.append(missionary.id)
                    continue
                missionary = session.get(Missionary, missionary_id)
                if action == "restore":
                    missionary.status = "ACTIVE"
                    missionary.deleted_at = None
                    if missionary.folder_path:
                        restored_folders.append(
                            (missionary.id, missionary.folder_path)
                        )
                    restored += 1
                changed = self._changed_fields(missionary, row)
                old_profile = missionary.tracking_profile or "LEGAL"
                if not changed and action != "restore":
                    unchanged += 1
                    continue
                for field in UPDATE_FIELDS:
                    setattr(missionary, field, row[field])
                missionary.missionary_code = row["missionary_code"]
                missionary.tracking_profile = profile
                if profile == "PERUVIAN_DNI":
                    missionary.current_stage = "DNI"
                    (
                        session.query(SecretaryTask)
                        .filter(
                            SecretaryTask.missionary_id == missionary.id,
                            SecretaryTask.automation_key.like("gvm:%"),
                            SecretaryTask.status.in_(
                                ("OPEN", "READY", "WAITING")
                            ),
                        )
                        .update(
                            {
                                SecretaryTask.status: "ARCHIVED",
                                SecretaryTask.automation_status_reason:
                                    "Missionary changed to Peruvian DNI tracking.",
                            },
                            synchronize_session=False,
                        )
                    )
                elif old_profile == "PERUVIAN_DNI":
                    missionary.current_stage = "INTERPOL"
                    legal_workflow_ids.append(missionary.id)
                updated += 1
                affected_ids.append(missionary.id)

            audit.status = "APPLIED"
            audit.applying_device = str(applying_device or "")
            audit.completed_at = datetime.now(timezone.utc)
            result = {
                "created": created, "updated": updated, "unchanged_count": unchanged,
                "restored": restored, "skipped_count": len(classified["skipped"]),
                "affected_missionary_ids": affected_ids,
            }
            audit.summary_json = json.dumps({
                key: value for key, value in result.items()
                if key != "affected_missionary_ids"
            })
            session.commit()
        except IntegrityError as exc:
            session.rollback()
            raise DynamicsRosterError(
                "A Missionary ID was assigned by another import. Preview the roster again."
            ) from exc
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

        for missionary_id, folder_path in restored_folders:
            try:
                destination = OneDriveService().restore_missionary_folder(
                    folder_path
                )
                restore_session = SessionLocal()
                try:
                    missionary = restore_session.get(Missionary, missionary_id)
                    if missionary is not None:
                        missionary.folder_path = str(destination)
                        restore_session.commit()
                finally:
                    restore_session.close()
            except Exception:
                logger.exception(
                    "Roster restored missionary %s but folder restore failed",
                    missionary_id,
                )
                result["folder_restore_warning"] = True

        for missionary_id in set(legal_workflow_ids):
            try:
                WorkflowService().initialize_workflows(missionary_id)
            except Exception:
                logger.exception(
                    "Roster applied but workflow initialization failed for %s",
                    missionary_id,
                )
                result["workflow_initialization_warning"] = True
        try:
            self._reconcile_automation()
        except Exception:
            logger.exception("Roster applied but automatic task reconciliation failed")
            result["automation_reconciliation_warning"] = True
        return {
            "preview_id": preview_id, "file_sha256": digest,
            "filename": Path(filename).name, **result,
        }

    def last_import(self):
        session = SessionLocal()
        try:
            audit = (
                session.query(DynamicsRosterImport)
                .filter_by(status="APPLIED")
                .order_by(DynamicsRosterImport.completed_at.desc())
                .first()
            )
            if audit is None:
                return None
            return {
                "preview_id": audit.preview_id, "filename": audit.filename,
                "filename_timestamp": audit.filename_timestamp,
                "file_sha256": audit.file_sha256,
                "dynamics_modified_at": audit.dynamics_modified_at,
                "counts": json.loads(audit.summary_json or "{}"),
                "applying_device": audit.applying_device,
                "completed_at": audit.completed_at,
            }
        finally:
            session.close()

    @staticmethod
    def _reconcile_automation():
        from services.process_automation_service import ProcessAutomationService
        # Reconcile date-keyed tasks immediately so changed roster dates and
        # dormant/profile transitions cannot leave stale Dashboard work.
        return object.__getattribute__(ProcessAutomationService(), "run")()

    @staticmethod
    def _filename_timestamp(filename):
        match = re.search(
            r"(\d{1,2}-[A-Za-z]{3}-\d{2} \d{1,2}-\d{2}-\d{2} [AP]M)",
            Path(filename).stem,
        )
        return match.group(1) if match else None

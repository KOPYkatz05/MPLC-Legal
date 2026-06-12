from services.missionary_service import missionary_display_id
from utils.i18n import tr
from utils.logger import logger


class ExportService:

    def _headers(self):
        return [
            "ID",
            "Full Name",
            "Nationality",
            "Passport Number",
            "Current Stage",
            "Arrival Date",
            "Visa Expiration",
            tr("export_passport_exp"),
            "Residency Expiration",
            "Carnet Issue Date",
            "Cancelación Date",
            tr("export_interpol_appt"),
            tr("export_biometric_appt"),
            tr("export_pickup_appt"),
            "Notes",
        ]

    def export_missionaries_to_excel(
        self,
        missionaries,
        output_path,
        columns=None,
    ):
        try:
            import openpyxl
            from openpyxl.styles import (
                Font,
                PatternFill,
                Alignment,
                Border,
                Side,
            )

            export_columns = self._export_columns(columns)
            headers = [
                column["label"]
                for column in export_columns
            ]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Missionaries"

            header_fill = PatternFill(
                start_color="1D4ED8",
                end_color="1D4ED8",
                fill_type="solid",
            )

            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=11,
            )

            header_border = Border(
                bottom=Side(
                    border_style="thin",
                    color="FFFFFF",
                ),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            ws.row_dimensions[1].height = 30

            alt_fill = PatternFill(
                start_color="EFF6FF",
                end_color="EFF6FF",
                fill_type="solid",
            )

            for i, m in enumerate(missionaries, 2):
                values = [
                    missionary_display_id(m),
                    m.full_name or "",
                    m.nationality or "",
                    m.passport_number or "",
                    m.current_stage or "",
                    fmt_date(m.arrival_date),
                    fmt_date(m.visa_expiration),
                    fmt_date(m.passport_expiration),
                    fmt_date(m.residency_expiration),
                    fmt_date(m.carnet_issue_date),
                    fmt_date(m.cancelacion_date),
                    fmt_date(m.interpol_appointment_date),
                    fmt_date(m.biometric_appointment_date),
                    fmt_date(m.pickup_appointment_date),
                    m.notes or "",
                ]

                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.alignment = Alignment(vertical="center")
                    if i % 2 == 0:
                        cell.fill = alt_fill

                ws.row_dimensions[i].height = 22

            col_widths = [
                12, 30, 16, 18, 22, 14, 16, 16, 20,
                16, 16, 16, 16, 16, 40,
            ]

            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[
                    ws.cell(row=1, column=col).column_letter
                ].width = column["width"]

            ws.freeze_panes = "A2"
            wb.save(output_path)

            logger.info(
                f"Exported {len(missionaries)} "
                f"missionaries to {output_path}"
            )

            return True

        except Exception:
            logger.exception(
                "Failed to export missionaries to Excel"
            )

            return False

    def _export_columns(self, columns):
        if columns:
            return [
                {
                    "label": column.label,
                    "getter": column.getter,
                    "width": self._excel_width(
                        getattr(column, "default_width", 120)
                    ),
                }
                for column in columns
            ]

        return self._legacy_export_columns()

    @staticmethod
    def _fmt_date(value):
        if not value:
            return ""

        return value.strftime("%d/%m/%Y")

    @staticmethod
    def _excel_width(pixel_width):
        return max(12, min(48, round(pixel_width / 8)))

    def _legacy_export_columns(self):
        return [
            {
                "label": header,
                "getter": getter,
                "width": width,
            }
            for header, getter, width in [
                ("ID", missionary_display_id, 12),
                ("Full Name", lambda m: m.full_name or "", 30),
                ("Nationality", lambda m: m.nationality or "", 16),
                ("Passport Number", lambda m: m.passport_number or "", 18),
                ("Current Stage", lambda m: m.current_stage or "", 22),
                (
                    "Arrival Date",
                    lambda m: self._fmt_date(m.arrival_date),
                    14,
                ),
                (
                    "Visa Expiration",
                    lambda m: self._fmt_date(m.visa_expiration),
                    16,
                ),
                (
                    tr("export_passport_exp"),
                    lambda m: self._fmt_date(m.passport_expiration),
                    16,
                ),
                (
                    "Residency Expiration",
                    lambda m: self._fmt_date(m.residency_expiration),
                    20,
                ),
                (
                    "Prorroga Expiration",
                    lambda m: self._fmt_date(m.prorroga_expiration),
                    20,
                ),
                (
                    "Carnet Issue Date",
                    lambda m: self._fmt_date(m.carnet_issue_date),
                    16,
                ),
                (
                    "Cancelacion Date",
                    lambda m: self._fmt_date(m.cancelacion_date),
                    16,
                ),
                (
                    tr("export_interpol_appt"),
                    lambda m: self._fmt_date(
                        m.interpol_appointment_date
                    ),
                    16,
                ),
                (
                    tr("export_biometric_appt"),
                    lambda m: self._fmt_date(
                        m.biometric_appointment_date
                    ),
                    16,
                ),
                (
                    tr("export_pickup_appt"),
                    lambda m: self._fmt_date(
                        m.pickup_appointment_date
                    ),
                    16,
                ),
                ("Notes", lambda m: m.notes or "", 40),
            ]
        ]

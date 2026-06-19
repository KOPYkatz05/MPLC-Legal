import zipfile

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database.db import Base
from database.models.missionary import Missionary
from database.models.secretary_work import MissionaryGroup, MissionaryGroupMember
from services import group_package_export_service as service_module
from services.group_package_export_service import GroupPackageExportService


def _testing_session(monkeypatch):
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    monkeypatch.setattr(service_module, "SessionLocal", TestingSession)
    return TestingSession


def test_full_group_export_creates_zip_with_excel_and_document_folders(
    monkeypatch,
    tmp_path,
):
    TestingSession = _testing_session(monkeypatch)
    elder_folder = tmp_path / "elder_docs"
    sister_folder = tmp_path / "sister_docs"
    elder_stage = elder_folder / "INTERPOL"
    sister_stage = sister_folder / "GENERAL"
    legacy_raw = elder_folder / "RAW_SCANS"
    legacy_processed = elder_folder / "OCR_PROCESSED"
    elder_stage.mkdir(parents=True)
    sister_stage.mkdir(parents=True)
    legacy_raw.mkdir()
    legacy_processed.mkdir()
    (elder_stage / "passport.pdf").write_text("passport", encoding="utf-8")
    (sister_stage / "tam.pdf").write_text("tam", encoding="utf-8")
    (legacy_raw / "raw.png").write_text("raw", encoding="utf-8")
    (legacy_processed / "processed.png").write_text(
        "processed",
        encoding="utf-8",
    )

    session = TestingSession()
    group = MissionaryGroup(name="June Arrivals")
    elder = Missionary(
        missionary_code="1042",
        full_name="Elder Smith",
        status="ACTIVE",
        tramite_usuario="elder-user",
        tramite_contrasena="elder-secret",
        folder_path=str(elder_folder),
    )
    sister = Missionary(
        missionary_code="1043",
        full_name="Sister Garcia",
        status="ACTIVE",
        tramite_usuario="sister-user",
        tramite_contrasena="sister-secret",
        folder_path=str(sister_folder),
    )
    session.add_all([group, elder, sister])
    session.flush()
    session.add_all(
        [
            MissionaryGroupMember(
                group_id=group.id,
                missionary_id=elder.id,
            ),
            MissionaryGroupMember(
                group_id=group.id,
                missionary_id=sister.id,
            ),
        ]
    )
    session.commit()
    group_id = group.id
    session.close()

    output_zip = tmp_path / "june_export.zip"

    result = GroupPackageExportService().export_group_package(
        group_id,
        output_zip,
    )

    assert result.group_name == "June Arrivals"
    assert result.missionary_count == 2
    assert result.skipped_folders == []

    with zipfile.ZipFile(output_zip) as archive:
        names = set(archive.namelist())
        assert "missionaries.xlsx" in names
        assert "documents/1042 - Elder Smith/INTERPOL/passport.pdf" in names
        assert "documents/1043 - Sister Garcia/GENERAL/tam.pdf" in names
        assert "documents/1042 - Elder Smith/RAW_SCANS/raw.png" not in names
        assert (
            "documents/1042 - Elder Smith/OCR_PROCESSED/processed.png"
            not in names
        )
        archive.extract("missionaries.xlsx", tmp_path)

    import openpyxl

    workbook = openpyxl.load_workbook(tmp_path / "missionaries.xlsx")
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    values = {
        sheet.cell(row=row, column=headers.index("Tramite Usuario") + 1).value
        for row in range(2, sheet.max_row + 1)
    }
    passwords = {
        sheet.cell(row=row, column=headers.index("Tramite Contrasena") + 1).value
        for row in range(2, sheet.max_row + 1)
    }

    assert values == {"elder-user", "sister-user"}
    assert passwords == {"elder-secret", "sister-secret"}


def test_full_group_export_reports_missing_document_folder(monkeypatch, tmp_path):
    TestingSession = _testing_session(monkeypatch)

    session = TestingSession()
    group = MissionaryGroup(name="Missing Docs")
    missionary = Missionary(
        missionary_code="2001",
        full_name="No Folder",
        status="ACTIVE",
        folder_path=str(tmp_path / "does_not_exist"),
    )
    session.add_all([group, missionary])
    session.flush()
    session.add(
        MissionaryGroupMember(
            group_id=group.id,
            missionary_id=missionary.id,
        )
    )
    session.commit()
    group_id = group.id
    session.close()

    output_zip = tmp_path / "missing_docs.zip"

    result = GroupPackageExportService().export_group_package(
        group_id,
        output_zip,
    )

    assert result.skipped_folders == ["No Folder"]
    with zipfile.ZipFile(output_zip) as archive:
        assert "missionaries.xlsx" in archive.namelist()
